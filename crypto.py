import os
import json
import ast
import pandas as pd
import argparse
import time
import requests
import hmac
import math
import html
import hashlib
from tabulate import tabulate
from urllib.parse import urlencode
from typing import Optional, List
from datetime import datetime, date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv

load_dotenv()

try:
    # SDK 1.x
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    def chat(model, messages):
        return client.chat.completions.create(model=model, messages=messages)
except ImportError:
    # SDK 0.28 (legacy)
    import openai
    openai.api_key = os.getenv("OPENAI_API_KEY")
    def chat(model, messages):
        return openai.ChatCompletion.create(model=model, messages=messages)

# Uso:
resp = chat("gpt-4o-mini", [{"role":"user","content":"Hola"}])

API_KEY = os.getenv("BINANCE_API_KEY")
API_SECRET = os.getenv("BINANCE_API_SECRET")

BASE_URL = 'https://api.binance.com'

EXCEL_FILE = "crypto.xlsx"
CRYPTO_SYMBOLS = ["BTC", "ETH", "SOL", "BCH", "BNB", "USDT"]
BINANCE_IDS = {
    "BTC": "bitcoin",
    "ETH": "ethereum",
    "SOL": "solana",
    "BCH": "bitcoin-cash",
    "BNB": "binancecoin",
    "USDT": "tether",
}
COLUMN_COLORS = {
    "BTC": "FFA500",  # naranja
    "ETH": "00BFFF",  # celeste
    "SOL": "8A2BE2",  # violeta
    "BCH": "008000",   # verde
    "BNB": "F0B90B",   # amarillo
    "USDT": "06402B",   # verde oscuro
}
HEADERS = [
    "Fecha de Compra", "Cripto", "Cantidad", "Precio Compra (USD)",
    "Valor Actual (USD)", "Wallet", "¿Stakeado? (Sí/No)"
]

# CLI args
parser = argparse.ArgumentParser()
for symbol in CRYPTO_SYMBOLS:
    parser.add_argument(f"--{symbol.lower()}", type=float, default=None)
    parser.add_argument(f"--{symbol.lower()}_wallet", type=str, default="")
    parser.add_argument(
        f"--{symbol.lower()}_date",
        type=lambda s: datetime.strptime(s, "%Y-%m-%d").date(),
        default=date.today(),
        help="Fecha en formato YYYY-MM-DD (por defecto: hoy)"
    )

# Agregar argumento opcional tipo bandera (booleano)
parser.add_argument('--binance', action='store_true', help='Ejecutar lógica para Binance')
parser.add_argument('--top10', action='store_true', help='Ejecutar lógica para Top 10')
parser.add_argument('--tendencia', action='store_true', help='Ejecutar lógica para tendencia en seis meses')
parser.add_argument('--bajas', action='store_true', help='Ejecutar lógica para las crypto bajas con posibilidad de crecimiento')
parser.add_argument('--telegram', action='store_true', help='Enviar un mensaje al bot de telegram con la tabla de totales del excel')
parser.add_argument('--usdt_tax', type=float, default=1)

args = parser.parse_args()
montos_usd = {s: getattr(args, s.lower()) for s in CRYPTO_SYMBOLS}
wallets = {s: getattr(args, f"{s.lower()}_wallet") for s in CRYPTO_SYMBOLS}
dates = {s: getattr(args, f"{s.lower()}_date") for s in CRYPTO_SYMBOLS}
binance = args.binance
top10 = args.top10
bajas = args.bajas
tendencia = args.tendencia
telegram = args.telegram
usdt_tax = args.usdt_tax

# Obtener precios actuales todos juntos
def get_all_prices(vs="usd"):
    url = "https://api.coingecko.com/api/v3/simple/price"
    params = {
        "ids": ",".join(BINANCE_IDS.values()),
        "vs_currencies": vs
    }
    try:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        # Mapea a {'BTC': price, ..., 'USDT': price} con None si falta
        return {
            sym: (data.get(cid) or {}).get(vs)
            for sym, cid in BINANCE_IDS.items()
        }
    except Exception:
        # Si algo falla, todos a None
        return {sym: None for sym in BINANCE_IDS}


def get_historical_prices(symbol):
    try:
        binance_symbol = f"{symbol.upper()}USDT"
        intervalo = "1d"
        limite = 180

        url = f"{BASE_URL}/api/v3/klines"
        params = {
            "symbol": binance_symbol,
            "interval": intervalo,
            "limit": limite
        }

        response = requests.get(url, params=params, headers={"X-MBX-APIKEY": API_KEY})
        response.raise_for_status()
        data = response.json()

        if not data:
            print(f"⚠️ Binance devolvió una respuesta vacía para {symbol}")
            return pd.DataFrame()

        # Convertir la respuesta en DataFrame
        df = pd.DataFrame(data, columns=[
            "open_time", "open", "high", "low", "close", "volume",
            "close_time", "quote_asset_volume", "number_of_trades",
            "taker_buy_base_volume", "taker_buy_quote_volume", "ignore"
        ])

        df["Fecha"] = pd.to_datetime(df["open_time"], unit="ms").dt.date
        df[symbol] = df["close"].astype(float)
        df = df[["Fecha", symbol]]
        return df

    except Exception as e:
        print(f"❌ Error al obtener históricos para {symbol}: {e}")
        return pd.DataFrame()


def get_price_on_date(symbol, fecha: date = date.today()):
    """
    Intenta obtener el precio de cierre para la fecha dada.
    Si la fecha es hoy y no hay datos, reintenta con el día anterior.
    """
    try:
        if symbol.upper() == "USDTUSDT":
            return 1 / usdt_tax  # USDT siempre es 1 USD

        def fetch_price(fecha_objetivo):
            start_ts = int(datetime.combine(fecha_objetivo, datetime.min.time()).timestamp() * 1000)
            end_ts = int(datetime.combine(fecha_objetivo + timedelta(days=1), datetime.min.time()).timestamp() * 1000)
            url = f"{BASE_URL}/api/v3/klines"
            params = {
                "symbol": symbol.upper(),
                "interval": "1d",
                "startTime": start_ts,
                "endTime": end_ts,
                "limit": 1
            }
            response = requests.get(url, params=params, headers={"X-MBX-APIKEY": API_KEY})
            response.raise_for_status()
            data = response.json()
            return float(data[0][4]) if data else None

        precio = fetch_price(fecha)
        if precio is not None:
            return precio

        # Si no encontró datos y la fecha es hoy, intentar con el día anterior
        if fecha == date.today():
            print(f"🔁 No se encontró precio para {symbol} en {fecha}, intentando con el día anterior...")
            return fetch_price(fecha - timedelta(days=1))

        print(f"⚠️ No se encontró precio para {symbol} en {fecha}")
        return None

    except Exception as e:
        print(f"❌ Error al obtener precio para {symbol} en {fecha}: {e}")
        return None


# Cargar o crear archivo
if os.path.exists(EXCEL_FILE):
    book = load_workbook(EXCEL_FILE)
else:
    book = Workbook()
    book.remove(book.active)

# Obtener precios actuales
all_prices = get_all_prices()
print(all_prices)

# Cryptos con montos actualizados
currency_updates = []

# --- Helpers para evitar que se "estire" por formatos y fórmulas ---
# Columnas de datos reales (A..G según tus HEADERS)
DATA_COLS = range(1, 8)  # 1=A, 7=G

def last_data_row(sheet: Worksheet, header_row: int = 1, data_cols=DATA_COLS) -> int:
    """Devuelve la última fila que tiene algún valor REAL en las columnas de datos."""
    min_c, max_c = min(data_cols), max(data_cols)
    for r in range(sheet.max_row, header_row, -1):
        for c in range(min_c, max_c + 1):
            v = sheet.cell(row=r, column=c).value
            if v not in (None, ""):
                return r
    return header_row

def write_df_after_last(sheet: Worksheet, df: pd.DataFrame, header_row: int = 1, data_cols=DATA_COLS) -> int:
    """Escribe el df empezando inmediatamente después de la última fila con datos."""
    start = last_data_row(sheet, header_row, data_cols) + 1
    r_idx = start
    for r in df.itertuples(index=False):
        for c_idx, val in enumerate(r, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=val)
        r_idx += 1
    return r_idx - 1  # última fila escrita

# --- Tu loop principal, usando las helpers ---
for symbol in CRYPTO_SYMBOLS:
    try:
        monto = montos_usd[symbol]
        if not monto:
            monto = 0  # si no hay monto, no seguimos

        date_crypto = dates[symbol] or date.today()
        binance_symbol = f"{symbol}USDT"
        precio_historico = get_price_on_date(binance_symbol, date_crypto)

        if symbol == "USDT":
            precio_historico = 1

        if not precio_historico:
            print(f"⚠️ No se pudo obtener el precio histórico para {symbol} en {date_crypto}")
            continue  # no seguimos si no hay precio

        cantidad = round(monto / precio_historico, 8)
        precio_compra = round(monto * usdt_tax, 2)
        valor_actual = round(precio_historico, 2)
        wallet = wallets[symbol] or ""

        fila = {
            "Fecha de Compra": date_crypto,
            "Cripto": symbol,
            "Cantidad": cantidad,
            "Precio Compra (USD)": precio_compra,
            "Valor Actual (USD)": valor_actual,
            "Wallet": wallet,
            "¿Stakeado? (Sí/No)": "No"
        }

        currency_updates.append(symbol)
        df = pd.DataFrame([fila])

        # Crear solapa si no existe
        if symbol in book.sheetnames:
            sheet = book[symbol]
        else:
            sheet = book.create_sheet(symbol)
            # encabezados
            for r in dataframe_to_rows(pd.DataFrame(columns=HEADERS), index=False, header=True):
                sheet.append(r)
            # estilos de encabezado
            for col_index, header in enumerate(HEADERS, start=1):
                cell = sheet.cell(row=1, column=col_index)
                cell.fill = PatternFill(
                    start_color=COLUMN_COLORS[symbol],
                    end_color=COLUMN_COLORS[symbol],
                    fill_type="solid"
                )
                cell.font = Font(color="FFFFFF", bold=True)

        # Si no hay cantidad (puede pasar por alguna división), no tiene sentido registrar
        if cantidad:
            # Agregar fila inmediatamente después de la última con datos reales
            write_df_after_last(sheet, df)
            print(f"✔ Entrada registrada para {symbol}")

        # Calcular totales
        headers = [cell.value for cell in sheet[1]]
        try:
            cantidad_idx = headers.index("Cantidad") + 1
            precio_idx = headers.index("Precio Compra (USD)") + 1
        except ValueError:
            continue

        resumen_row = 2
        col_tot_cantidad = 8  # columna H
        col_tot_precio = 9    # columna I

        # Etiquetas Totales
        cellTotal = sheet.cell(row=1, column=col_tot_cantidad, value="Total Crypto")
        cellPrecio = sheet.cell(row=1, column=col_tot_precio, value="Total USD")
        bgCell = PatternFill(
            start_color=COLUMN_COLORS[symbol],
            end_color=COLUMN_COLORS[symbol],
            fill_type="solid"
        )
        colorCell = Font(color="FFFFFF", bold=True)
        cellTotal.fill = cellPrecio.fill = bgCell
        cellTotal.font = cellPrecio.font = colorCell

        # Usa la última fila con datos reales, no max_row (que puede estar inflado por formatos)
        ultima_dato = last_data_row(sheet, header_row=1, data_cols=DATA_COLS)

        suma_cantidad = (
            f"=SUM({get_column_letter(cantidad_idx)}2:{get_column_letter(cantidad_idx)}{ultima_dato})"
            if cantidad_idx else "0"
        )
        suma_precio = (
            f"=SUM({get_column_letter(precio_idx)}2:{get_column_letter(precio_idx)}{ultima_dato})"
            if precio_idx else "0"
        )
        sheet.cell(row=resumen_row, column=col_tot_cantidad, value=suma_cantidad)
        sheet.cell(row=resumen_row, column=col_tot_precio, value=suma_precio)

    except Exception as e:
        print(f"⚠️ Error procesando {symbol}: {e}")
        continue

# Crear hoja solo si no existe (se regenera cada vez)
if "Totales" in book.sheetnames:
    del book["Totales"]

sheet = book.create_sheet("Totales")

# Encabezados + columnas de totales
headers = [
    "Cripto",
    "Cantidad Total",
    "Precio Compra Total (USD)",
    "Cantidad Actual",
    "Ganancia/Perdida (USD)",
    "Total Compra",   # suma de toda la col "Precio Compra Total (USD)"
    "Total Actual",   # suma de toda la col "Cantidad Actual"
    "Total Ganancia"  # suma de toda la col "Ganancia/Perdida (USD)"
]
sheet.append(headers)

# Estilo de encabezados
for col_index, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=col_index)
    cell.fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
    cell.font = Font(color="FFFFFF", bold=True)

# Reservamos la fila 2 para las sumas (debajo del encabezado)
sheet.insert_rows(2)

# Insertar nueva fila para cada cripto con datos (comienzan en fila 3)
for symbol in CRYPTO_SYMBOLS:
    if symbol not in currency_updates:
        continue

    cantidad_total = 0.0
    precio_total = 0.0

    if symbol in book.sheetnames:
        sheet_symbol = book[symbol]
        for row in sheet_symbol.iter_rows(min_row=2, values_only=True):
            try:
                # Índices según tus HEADERS en cada hoja de cripto:
                # 0 Fecha de Compra, 1 Cripto, 2 Cantidad, 3 Precio Compra (USD), 4 Valor Actual (USD), ...
                cantidad_total += float(row[2]) if row[2] not in (None, "", "N/A") else 0.0
                precio_total   += float(row[3]) if row[3] not in (None, "", "N/A") else 0.0
            except (ValueError, IndexError) as e:
                print(f"⚠️ Error procesando fila en {symbol}: {row} — {e}")
                continue

    precio_actual = all_prices.get(symbol)
    cantidad_actual = round(cantidad_total * precio_actual, 2) if cantidad_total and precio_actual else ""
    cantidad_actual = round(cantidad_total * precio_actual, 2) if cantidad_total and precio_actual else ""
    ganancia_perdida = round(cantidad_actual - precio_total, 2) if cantidad_actual != "" else ""

    sheet.append([symbol, cantidad_total, precio_total, cantidad_actual, ganancia_perdida, "", "", ""])

# Ahora que ya están las filas, escribimos las SUMAS en la fila 2
last_row = sheet.max_row

# Columnas (1-indexed) según 'headers'
col_precio_compra = headers.index("Precio Compra Total (USD)") + 1  # C = 3
col_cantidad_actual = headers.index("Cantidad Actual") + 1           # D = 4
col_ganancia = headers.index("Ganancia/Perdida (USD)") + 1          # E = 5
col_total_compra = headers.index("Total Compra") + 1                # F = 6
col_total_actual = headers.index("Total Actual") + 1                # G = 7
col_total_ganancia = headers.index("Total Ganancia") + 1            # H = 8

# Si hay al menos una fila de datos (fila 3 en adelante), ponemos las fórmulas; si no, "0"
if last_row >= 3:
    rango_compra   = f"{get_column_letter(col_precio_compra)}2:{get_column_letter(col_precio_compra)}{last_row}"
    rango_actual   = f"{get_column_letter(col_cantidad_actual)}2:{get_column_letter(col_cantidad_actual)}{last_row}"
    rango_ganancia = f"{get_column_letter(col_ganancia)}2:{get_column_letter(col_ganancia)}{last_row}"

    sheet.cell(row=2, column=col_total_compra,   value=f"=SUM({rango_compra})")
    sheet.cell(row=2, column=col_total_actual,   value=f"=SUM({rango_actual})")
    sheet.cell(row=2, column=col_total_ganancia, value=f"=SUM({rango_ganancia})")
else:
    sheet.cell(row=2, column=col_total_compra,   value="0")
    sheet.cell(row=2, column=col_total_actual,   value="0")
    sheet.cell(row=2, column=col_total_ganancia, value="0")

for col in (col_total_compra, col_total_actual, col_total_ganancia):
    c = sheet.cell(row=2, column=col)
    c.font = Font(bold=True)

top10_telegram = ""
bajas_telegram = ""

if top10:
     # Crear hoja "top10"
    if "Top 10 Cripto del momento" in book.sheetnames:
        del book["Top 10 Cripto del momento"]

    sheet = book.create_sheet("Top 10 Cripto del momento")

    def get_top10_crypto(max_retries=3, delay=2):
        prompt_messages = [
            {"role": "system", "content": (
                "I am aware that you are not a financial analyst, far from it."
                "I also know that no serious financial analyst would dare to predict the future or induce people to invest based on their recommendations."
                "However, I want you to gather the presumed opinions of some and tell me which 10 cryptos are presumed to be on the rise in the next six months."
            )},
            {"role": "user", "content": (
                "Your only response should be a valid Python list, for example ['BTC', 'ETH', 'SOL', ... 'reflection...'], with no explanations, no additional text."
                "Return them ordered, first the one expected to rise the most and so on."
                "The last item in the array should not be one of the rising cryptos, but a random suggestion of a good trading practice to improve, considering general investment knowledge, but also practical suggestions for Binance tools."
            )}
        ]


        for attempt in range(1, max_retries + 1):
            try:
                response = client.chat.completions.create(model="gpt-4",
                messages=prompt_messages)
                top10_text = response.choices[0].message.content

                # Intentar interpretar como lista Python
                top10 = ast.literal_eval(top10_text)

                # Validar que sea lista con al menos un string
                if isinstance(top10, list) and all(isinstance(x, str) for x in top10) and len(top10) > 0:
                    return top10
                else:
                    print("⚠️ La respuesta no es una lista válida de strings no vacía.")

            except Exception as e:
                print(f"⚠️ Error al interpretar la respuesta (intento {attempt}): {e}")

            if attempt < max_retries:
                print("🔁 Reintentando...")
                time.sleep(delay)

        # Si todo falla, retornar None
        return None

    # Llamar a la función
    top10 = get_top10_crypto()

    if top10:
        print("✅ Top 10 obtenido: ", top10)
        top10_telegram = "✅ Top 3 obtenido: " + ", ".join(top10[:3])
        df_top10 = pd.DataFrame({"Top 10 Cripto del momento": top10})

        for r in dataframe_to_rows(df_top10, index=False, header=True):
            sheet.append(r)
            
         # --- Función para buscar si ya existe una fila con cierto texto ---
        def find_row_index(sheet, target_text):
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and row[0] == target_text:
                    return i
            return None

        # --- Insertar o sobrescribir Top 10 ---
        top10_header = "Top 10 Cripto del momento"
        top10_row_idx = find_row_index(sheet, top10_header)

        if top10_row_idx:
            for i, r in enumerate(dataframe_to_rows(df_top10, index=False, header=True)):
                for j, value in enumerate(r, start=1):
                    sheet.cell(row=top10_row_idx + i, column=j, value=value)
        else:
            for r in dataframe_to_rows(df_top10, index=False, header=True):
                sheet.append(r)
    else:
        print("❌ No se pudo obtener un Top 10 válido.")

if tendencia:
    # Crear hoja "tendencia"
    if "Tendencia de mis criptos" in book.sheetnames:
        del book["Tendencia de mis criptos"]

    sheet = book.create_sheet("Tendencia de mis criptos")

    # Línea vacía + encabezado de tabla real
    sheet.append([])
    sheet.append(["Histórico de precios (USD) - Últimos 6 meses"])

    # Obtener históricos de precios
    historical_symbols = ["BTC", "ETH", "SOL", "BCH", "BNB"]
    df_hist = pd.DataFrame()
    for symbol in historical_symbols:
        df_temp = get_historical_prices(symbol)
        if df_temp.empty or "Fecha" not in df_temp.columns:
            print(f"⚠️ No se pudieron obtener datos históricos para {symbol}")
            continue
        if df_hist.empty:
            df_hist = df_temp
        else:
            df_hist = pd.merge(df_hist, df_temp, on="Fecha", how="outer")

    if not df_hist.empty and "Fecha" in df_hist.columns:
        df_hist = df_hist.sort_values("Fecha").ffill().tail(180)
    else:
        print("⚠️ No se pudo generar la tabla de históricos. df_hist está vacío.")

    for col in historical_symbols:
        if col not in df_hist.columns:
            df_hist[col] = None

    # Agregar tabla real
    start_row_real = sheet.max_row + 1
    for r in dataframe_to_rows(df_hist, index=False, header=True):
        sheet.append(r)

    # Colorear encabezados reales
    header_row_real = sheet[start_row_real]
    for cell in header_row_real:
        col_name = cell.value
        if col_name in COLUMN_COLORS:
            cell.fill = PatternFill(start_color=COLUMN_COLORS[col_name], end_color=COLUMN_COLORS[col_name], fill_type="solid")
        else:
            cell.fill = PatternFill()  # sin color para "Fecha"

    # Gráfico real
    chart_real = LineChart()
    chart_real.title = "Evolución Absoluta (USD)"
    chart_real.y_axis.title = "Precio USD"
    chart_real.x_axis.title = "Fecha"

    end_row_real = sheet.max_row
    data_real = Reference(sheet, min_col=2, max_col=6, min_row=start_row_real, max_row=end_row_real)
    cats_real = Reference(sheet, min_col=1, min_row=start_row_real + 1, max_row=end_row_real)
    chart_real.add_data(data_real, titles_from_data=True)
    chart_real.set_categories(cats_real)

    for i, ser in enumerate(chart_real.series):
        ser.graphicalProperties.line.solidFill = list(COLUMN_COLORS.values())[i]

    sheet.add_chart(chart_real, f"G2")

    # Tabla normalizada
    sheet.append([])
    sheet.append(["Histórico de precios normalizados (Base 100)"])
    df_norm = df_hist.copy()

    if not df_norm.empty and df_norm.shape[0] > 0:
        try:
            df_norm.iloc[:, 1:] = df_norm.iloc[:, 1:].astype(float)
            base_values = df_norm.iloc[0, 1:].astype(float)
            df_norm.iloc[:, 1:] = df_norm.iloc[:, 1:].div(base_values).multiply(100)
        except Exception as e:
            print(f"⚠️ No se pudo normalizar el histórico: {e}")
            df_norm = pd.DataFrame()
    else:
        print("⚠️ No hay datos suficientes para normalizar el histórico.")
        df_norm = pd.DataFrame()


    start_row_norm = sheet.max_row + 1
    for r in dataframe_to_rows(df_norm, index=False, header=True):
        sheet.append(r)

    # Colorear encabezados normalizados
    header_row_norm = sheet[start_row_norm]
    for cell in header_row_norm:
        col_name = cell.value
        if col_name in COLUMN_COLORS:
            cell.fill = PatternFill(start_color=COLUMN_COLORS[col_name], end_color=COLUMN_COLORS[col_name], fill_type="solid")
        else:
            cell.fill = PatternFill()

    # Gráfico normalizado
    chart_norm = LineChart()
    chart_norm.title = "Evolución Relativa (Base 100)"
    chart_norm.y_axis.title = "Índice base 100"
    chart_norm.x_axis.title = "Fecha"

    end_row_norm = sheet.max_row
    data_norm = Reference(sheet, min_col=2, max_col=6, min_row=start_row_norm, max_row=end_row_norm)
    cats_norm = Reference(sheet, min_col=1, min_row=start_row_norm + 1, max_row=end_row_norm)
    chart_norm.add_data(data_norm, titles_from_data=True)
    chart_norm.set_categories(cats_norm)

    for i, ser in enumerate(chart_norm.series):
        ser.graphicalProperties.line.solidFill = list(COLUMN_COLORS.values())[i]

    sheet.add_chart(chart_norm, f"G30")

if bajas:
     # Crear hoja "top10"
    if "Top 10 Cripto bajas" in book.sheetnames:
        del book["Top 10 Cripto bajas"]

    sheet = book.create_sheet("Top 10 Cripto bajas")

    def get_top10_bajas_crypto(max_retries=3, delay=2):
        prompt_messages = [
            {"role": "system", "content": (
                "I am aware that you are not a financial analyst, far from it."
                "I also know that no serious financial analyst would dare to predict the future or induce people to invest based on their recommendations."
                "However, I want you to gather the opinions of the most prestigious analysts and tell me which 10 cryptos that start at a very low price are presumed to have potential to rise in the next six months, based on the seriousness of those behind them."
            )},
            {"role": "user", "content": (
                "Your only response should be a valid Python list, for example ['SOL', 'BNB', 'ABA'], with no explanations, no additional text."
                "Return them ordered, first the one with the most potential and so on."
            )}
        ]

        for attempt in range(1, max_retries + 1):
            try:
                response = client.chat.completions.create(model="gpt-4",
                messages=prompt_messages)
                top10_bajas_text = response.choices[0].message.content

                # Intentar interpretar como lista Python
                top10_bajas = ast.literal_eval(top10_bajas_text)

                # Validar que sea lista con al menos un string
                if isinstance(top10_bajas, list) and all(isinstance(x, str) for x in top10_bajas) and len(top10_bajas) > 0:
                    return top10_bajas
                else:
                    print("⚠️ La respuesta no es una lista válida de strings no vacía.")

            except Exception as e:
                print(f"⚠️ Error al interpretar la respuesta (intento {attempt}): {e}")

            if attempt < max_retries:
                print("🔁 Reintentando...")
                time.sleep(delay)

        # Si todo falla, retornar None
        return None

    # Llamar a la función
    top10_bajas = get_top10_bajas_crypto()

    if top10_bajas:
        print("✅ Top 10 de bajas obtenido:", top10_bajas)
        bajas_telegram = "✅ Top 3 bajas con potencial: " + ", ".join(top10[:3])
        df_top10_bajas = pd.DataFrame({"Top 10 Cripto a bajo costo con potencial ": top10_bajas})

        for r in dataframe_to_rows(df_top10_bajas, index=False, header=True):
            sheet.append(r)
        
         # --- Función para buscar si ya existe una fila con cierto texto ---
        def find_row_index(sheet, target_text):
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and row[0] == target_text:
                    return i
            return None

        # --- Insertar o sobrescribir Top 10 ---
        top10_bajas_header = "Top 10 Cripto a bajo costo con potencial"
        top10_bajas_row_idx = find_row_index(sheet, top10_bajas_header)

        if top10_bajas_row_idx:
            for i, r in enumerate(dataframe_to_rows(df_top10_bajas, index=False, header=True)):
                for j, value in enumerate(r, start=1):
                    sheet.cell(row=top10_bajas_row_idx + i, column=j, value=value)
        else:
            for r in dataframe_to_rows(df_top10_bajas, index=False, header=True):
                sheet.append(r)
    else:
        print("❌ No se pudo obtener un Top 10 válido.")

if binance:    
    def create_signature(params, secret):
        query_string = urlencode(params)
        return hmac.new(secret.encode(), query_string.encode(), hashlib.sha256).hexdigest()

    def signed_get(method, endpoint, api_key, api_secret, extra_params={}):
        timestamp = int(time.time() * 1000)
        params = {'timestamp': timestamp}
        params.update(extra_params)
        signature = create_signature(params, api_secret)
        params['signature'] = signature
        headers = {'X-MBX-APIKEY': api_key}
        url = f"{BASE_URL}{endpoint}"
        if method == "GET":
            response = requests.get(url, headers=headers, params=params, timeout=15)
        elif method == "POST":
            response = requests.post(url, headers=headers, params=params, timeout=15)
        else:
            raise ValueError("Método HTTP no soportado.")
        response.raise_for_status()
        return response.json()

    def get_usdt_prices_map(assets):
        """
        Devuelve un dict { 'BTC': 64321.12, 'ETH': 3210.55, 'USDT': 1.0, ... }.
        1) Intenta batch con symbols (JSON compacto, sin espacios).
        2) Si falla, pide todos los precios y filtra en local.
        """
        assets = [a for a in assets if a]  # limpia None/''

        # Atajo si solo hay USDT
        if set(assets) <= {"USDT"}:
            return {"USDT": 1.0}

        # 1) Intento batch con JSON compacto
        try:
            pairs = [f"{a}USDT" for a in assets if a != "USDT"]
            if pairs:
                url = f"{BASE_URL}/api/v3/ticker/price"
                # JSON sin espacios (Binance a veces rechaza espacios)
                symbols_param = json.dumps(pairs, separators=(',', ':'))
                resp = requests.get(url, params={"symbols": symbols_param}, timeout=15)
                resp.raise_for_status()
                data = resp.json()
                price_map = {
                    item["symbol"][:-4]: float(item["price"])
                    for item in data
                    if item.get("symbol", "").endswith("USDT") and "price" in item
                }
                price_map["USDT"] = 1.0
                return price_map
        except Exception as e:
            print(f"⚠️ Batch symbols falló, voy al fallback: {e}")

        # 2) Fallback: traer todos los precios y filtrar
        try:
            url = f"{BASE_URL}/api/v3/ticker/price"
            resp = requests.get(url, timeout=20)
            resp.raise_for_status()
            all_prices = resp.json()  # lista de {'symbol':'BTCUSDT','price':'...'}
            usdt_prices = {
                item["symbol"][:-4]: float(item["price"])
                for item in all_prices
                if item.get("symbol", "").endswith("USDT")
            }
            # Solo dejá lo que pedimos
            price_map = {a: usdt_prices.get(a) for a in assets if a != "USDT"}
            price_map["USDT"] = 1.0
            return price_map
        except Exception as e:
            print(f"⚠️ Error en fallback de precios: {e}")
            return {"USDT": 1.0}


    def get_spot():
        try:
            endpoint = "/api/v3/account"
            ts = int(time.time() * 1000)
            params = {"timestamp": ts, "recvWindow": 60000}
            data = signed_get("GET", endpoint, API_KEY, API_SECRET, params)

            # Filtramos solo balances con saldo > 0
            balances = [
                {
                    "Asset": b["asset"],
                    "Free": float(b["free"]),
                    "Locked": float(b["locked"]),
                    "Total": float(b["free"]) + float(b["locked"])
                }
                for b in data.get("balances", [])
                if float(b["free"]) + float(b["locked"]) > 0
            ]

            if not balances:
                return pd.DataFrame(columns=["Asset", "Free", "Locked", "Total", "Price (USDT)", "Total USD"])

            # Armamos el set de assets y traemos precios en un solo request
            assets = sorted({b["Asset"] for b in balances})
            prices_map = get_usdt_prices_map(assets)

            # Agregamos columnas de precio y total en USD
            for b in balances:
                asset = b["Asset"]
                price = prices_map.get(asset)
                b["Price (USDT)"] = price if price is not None else None
                b["Total USD"] = round(b["Total"] * price, 2) if price is not None else None

            # Orden de columnas prolijo
            cols = ["Asset", "Free", "Locked", "Total", "Price (USDT)", "Total USD"]
            return pd.DataFrame(balances, columns=cols)

        except Exception as e:
            print(f"❌ Error al obtener información de Binance: {e}")
            return pd.DataFrame(columns=["Asset", "Free", "Locked", "Total", "Price (USDT)", "Total USD"])

    def get_funding_balances():
        try:
            endpoint = "/sapi/v1/asset/get-funding-asset"
            timestamp = int(time.time() * 1000)

            params = {
                "timestamp": timestamp,
                "recvWindow": 60000,   # margen de reloj
            }
            query_string = urlencode(params)
            signature = hmac.new(API_SECRET.encode(), query_string.encode(), hashlib.sha256).hexdigest()
            params["signature"] = signature
            headers = {'X-MBX-APIKEY': API_KEY}
            url = f"{BASE_URL}{endpoint}"

            response = requests.post(url, headers=headers, params=params, timeout=20)
            response.raise_for_status()
            data = response.json()

            filtered = [
                {
                    'Asset': a['asset'],
                    'Free': float(a['free']),
                    'Locked': float(a.get('locked', 0)),
                    'Freeze': float(a.get("freeze")),
                    'Withdrawing': float(a.get("withdrawing")),
                    'Total': float(a['free']) + float(a.get('locked', 0)) + float(a.get("freeze", 0)) + float(a.get("withdrawing", 0))
                }
                for a in data if float(a['free']) + float(a.get('locked', 0)) > 0
            ]
            return pd.DataFrame(filtered)
        except Exception as e:
            print(f"❌ Error al obtener fondos: {e}")
            return pd.DataFrame()


    def write_df_to_sheet(wb, sheet_name, df):
         # Estilos
        black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
        white_bold_font = Font(color="FFFF00", bold=True)

        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        first = True
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)
            if first:
                last_row = ws.max_row  # justo se acaba de insertar
                for col in range(1, len(row) + 1):
                    cell = ws.cell(row=last_row, column=col)
                    cell.fill = black_fill
                    cell.font = white_bold_font
                first = False

    def main():
        df_orders = get_spot()
        df_funds = get_funding_balances()
        print(df_orders)
        print(df_funds)

        if not df_orders.empty:
            write_df_to_sheet(book, "Binance Spot", df_orders)

        if not df_funds.empty:
            write_df_to_sheet(book, "Binance Fondos", df_funds)

    if __name__ == "__main__":
        main()

# Guardar archivo
book.save(EXCEL_FILE)
print(f"✅ Archivo actualizado: {EXCEL_FILE}")

TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
DEFAULT_CHAT_ID = os.getenv("DEFAULT_CHAT_ID")
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
# Límite seguro para texto en Telegram (los bots aceptan ~4096 chars)
TG_LIMIT = 4000

def _read_sheet_values(excel_path: str, sheet_name: str) -> pd.DataFrame:

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    headers = list(rows[0])
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return df

def _format_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """2 dec para USD/totales, 8 dec para cantidades; resto texto."""
    out = df.copy()
    usd_like = { "Cripto",
        "Cantidad Total",
        "Precio Compra Total (USD)",
        "Cantidad Actual",
        "Ganancia/Perdida (USD)"}
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            if "Cantidad" in c and "USD" not in c:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.8f}".replace(",", "X").replace(".", ",").replace("X", "."))
            elif c in usd_like or "USD" in c:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            else:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        else:
            out[c] = out[c].astype(str)
    return out

def _send_html_pre(chat_id: str, title: str, body: str):
    if not TELEGRAM_TOKEN:
        raise RuntimeError("Falta TELEGRAM_TOKEN")
    payload = {
        "chat_id": chat_id or DEFAULT_CHAT_ID,
        "text": f"{html.escape(title)}\n<pre>{html.escape(body)}</pre>",
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    r = requests.post(TELEGRAM_API, json=payload, timeout=30)
    r.raise_for_status()

def send_telegram_table(
    sheet_name: str,
    *,
    excel_path: str,
    columns: Optional[List[str]] = None,
    top: Optional[int] = None,
    sort_by: Optional[str] = None,
    ascending: bool = False,
    rows_per_msg: int = 12,
    tablefmt: str = "github",
    title: Optional[str] = None,
    chat_id: Optional[str] = None,
) -> None:
    chat_id = chat_id or DEFAULT_CHAT_ID
    if not chat_id:
        raise RuntimeError("Falta DEFAULT_CHAT_ID/TELEGRAM_CHAT_ID")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No encuentro el Excel en: {excel_path}")

    # 1) Leer hoja
    df = _read_sheet_values(excel_path, sheet_name)

    # 2) Filtrar y ordenar
    if columns:
        cols = [c for c in columns if c in df.columns]
        df = df[cols] if cols else df
    if sort_by and sort_by in df.columns:
        df = df.sort_values(sort_by, ascending=ascending)

    if top is not None:
        df = df.head(top)

    # === NUEVO: calcular totales y diferencia sobre datos numéricos ===
    col_precio = "Precio Compra Total (USD)"
    col_cant   = "Cantidad Actual"

    precio_num = pd.to_numeric(df[col_precio], errors="coerce") if col_precio in df.columns else pd.Series(dtype=float)
    cant_num   = pd.to_numeric(df[col_cant], errors="coerce")   if col_cant   in df.columns else pd.Series(dtype=float)

    total_precio = float(precio_num.sum(skipna=True)) if len(precio_num) else 0.0
    total_cant   = float(cant_num.sum(skipna=True))   if len(cant_num)   else 0.0
    diferencia   = total_cant - total_precio

    # 5) Limpieza final para visual
    df = df.replace({None: pd.NA}).fillna("")

    # 4) Formatear números para la tabla
    df_fmt = _format_numeric_columns(df)
    n = len(df_fmt)
    pages = max(1, math.ceil(n / rows_per_msg)) if rows_per_msg else 1

    for p in range(pages):
        start = p * rows_per_msg
        end = None if not rows_per_msg else start + rows_per_msg
        page_df_fmt = df_fmt.iloc[start:end]  # 👈 usar el slice correcto

        body = tabulate(
            page_df_fmt,
            headers="keys",
            tablefmt=tablefmt,
            numalign="right",
            stralign="left",
            showindex=False
        )

        # === NUEVO: agregar totales al final de la ÚLTIMA página ===
        if p == pages - 1:

            footer_lines = [
                "",
                f"Total Precio Compra: {total_precio:,.0f}",
                f"Total Cantidad Actual: {total_cant:,.0f}",
                f"✅ Ganancia: {diferencia:,.0f}" if (diferencia > 0) else f"❌ Pérdida: {diferencia:,.0f}",
            ]
            body = body + "\n" + "\n".join(footer_lines) + "\n\n"
            
        if top10_telegram != "":
            body = body + "\n" + top10_telegram
        if bajas_telegram != "":
            body = body + "\n" + bajas_telegram

        # Recorte ultraseguro si excede límites
        if len(body) > 3900:
            body = "\n".join(line[:120] for line in body.splitlines())

        page_title = f"{title or sheet_name} ({p+1}/{pages})" if pages > 1 else (title or sheet_name)
        _send_html_pre(chat_id, page_title, body)
        
if telegram:
    headers_list = [
        "Cripto",
        "Cantidad Total",
        "Precio Compra Total (USD)",
        "Cantidad Actual",
        "Ganancia/Perdida (USD)"
    ]

    send_telegram_table(
        sheet_name="Totales",
        excel_path=EXCEL_FILE,        # usa tu constante existente
        columns=headers_list,
        top=20,
        sort_by="Total Ganancia",     # opcional: ordena por ganancia
        ascending=False,
        rows_per_msg=12,
        tablefmt="github",
        title="Totales",
        chat_id=DEFAULT_CHAT_ID
    )
