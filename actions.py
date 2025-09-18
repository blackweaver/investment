# https://polygon.io/dashboard/keys/4040a189-9e0d-4b0a-8c8e-5a1871da6bcf?values=flat-files
import os
import sys
import json
import ast
import requests
import argparse
import time
import math
import html
import random
from tabulate import tabulate
import pandas as pd
from typing import Dict, Any, List, Optional, Iterable
from datetime import datetime, date, timedelta
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import PatternFill, Font
from dotenv import load_dotenv
load_dotenv()

try:
    # SDK 1.x
    from openai import OpenAI
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
    def chat(model, messages):
        return client.chat.completions.create(model=model, messages=messages)
except ImportError:
    # SDK 0.28 (legacy)
    import openai
    openai.api_key = os.getenv("OPENAI_API_KEY")
    def chat(model, messages):
        return openai.ChatCompletion.create(model=model, messages=messages)

# Uso:
resp = chat("gpt-4o-mini", [{"role":"user","content":"Hola"}])

# Endpoints Polygon
SNAPSHOT_URL = "https://api.polygon.io/v2/snapshot/locale/us/markets/stocks/tickers"
PREV_URL_TMPL = "https://api.polygon.io/v2/aggs/ticker/{ticker}/prev"
REF_TICKER_URL_TMPL = "https://api.polygon.io/v3/reference/tickers/{ticker}"
HIST_URL_TMPL = "https://api.polygon.io/v2/aggs/ticker/{ticker}/range/1/day/{date}/{date}"
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
DEFAULT_CHAT_ID = os.getenv("DEFAULT_CHAT_ID")
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}/sendMessage"
# Límite seguro para texto en Telegram (los bots aceptan ~4096 chars)
TG_LIMIT = 4000

EXCEL_FILE = "actions.xlsx"
actions_sheet_name = "Acciones invertidas"
current_investment = {}
# Caché de nombres en memoria (podés persistirlo con shelve o json)
name_cache: Dict[str, str] = {}
# Companies con montos actualizados
currency_updates = []

# Variables de entorno (tal como pediste)
API_KEY_POLYGON = os.getenv("API_KEY_POLYGON")
TICKETS = os.getenv("TICKETS")

TICKETS_SYMBOLS = [t.strip() for t in TICKETS.split(",") if t.strip()]
def random_hex_color() -> str:
    """Devuelve un color random en formato hex (ej: 'A1B2C3')."""
    return f"{random.randint(0, 0xFFFFFF):06X}"

# Construir el diccionario dinámicamente
COLUMN_ACTION_COLORS = {
    symbol: random_hex_color() for symbol in TICKETS_SYMBOLS
}
HEADERS = [
    "Fecha de Compra", "Compañía", "Cantidad", "Precio Compra (USD)",
    "Valor Actual (USD)", "Plataforma"
]

# CLI args
parser = argparse.ArgumentParser(conflict_handler="resolve")
for symbol in TICKETS_SYMBOLS:
    parser.add_argument(f"--{symbol.lower()}", type=float, default=None)
    parser.add_argument(f"--{symbol.lower()}_platform", type=str, default="")
    parser.add_argument(
        f"--{symbol.lower()}_date",
        type=lambda s: datetime.strptime(s, "%Y-%m-%d").date(),
        default=None,
        help="Fecha en formato YYYY-MM-DD (por defecto: hoy)"
    )
        
parser.add_argument('--totales', action='store_true', help='Totales en hoja aparte')

parser.add_argument('--top10', action='store_true', help='Ejecutar lógica para Top 10')
parser.add_argument('--bajas', action='store_true', help='Ejecutar lógica para Bajas')
parser.add_argument('--telegram', action='store_true', help='Enviar un mensaje al bot de telegram con la tabla de totales del excel')
parser.add_argument(
    "--date",
    type=lambda s: datetime.strptime(s, "%Y-%m-%d").strftime("%Y-%m-%d"),
    default=None,
    help="Fecha en formato YYYY-MM-DD (por defecto: hoy)"
)

top10_telegram = ""
bajas_telegram = ""

args = parser.parse_args()
montos_usd = {s: getattr(args, s.lower()) for s in TICKETS_SYMBOLS}
wallets = {s: getattr(args, f"{s.lower()}_platform") for s in TICKETS_SYMBOLS}
dates = {s: getattr(args, f"{s.lower()}_date") for s in TICKETS_SYMBOLS}
top10 = args.top10
bajas = args.bajas
telegram = args.telegram
totales = args.totales

# Cargar o crear archivo
if os.path.exists(EXCEL_FILE):
    book = load_workbook(EXCEL_FILE)
else:
    book = Workbook()
    book.remove(book.active)
    
def get_price_on_date(symbol, fecha: date = date.today()):
    print(symbol, fecha)
    return None

DATA_COLS = range(1, 8)

def last_data_row(sheet: Worksheet, header_row: int = 1, data_cols=DATA_COLS) -> int:
    """Devuelve la última fila que tiene algún valor REAL en las columnas de datos."""
    min_c, max_c = min(data_cols), max(data_cols)
    for r in range(sheet.max_row, header_row, -1):
        for c in range(min_c, max_c + 1):
            v = sheet.cell(row=r, column=c).value
            if v not in (None, ""):
                return r
    return header_row

def write_df_after_last(sheet: Worksheet, df: pd.DataFrame, header_row: int = 1, data_cols=DATA_COLS) -> int:
    """Escribe el df empezando inmediatamente después de la última fila con datos."""
    start = last_data_row(sheet, header_row, data_cols) + 1
    r_idx = start
    for r in df.itertuples(index=False):
        for c_idx, val in enumerate(r, start=1):
            sheet.cell(row=r_idx, column=c_idx, value=val)
        r_idx += 1
    return r_idx - 1  # última fila escrita

# Top 10
if top10:
     # Crear hoja "top10"
    top10_header = "Top 10 acciones del momento"
    if top10_header in book.sheetnames:
        del book[top10_header]

    sheet = book.create_sheet(top10_header)

    def get_top10_crypto(max_retries=3, delay=2):
        prompt_messages = [
            {"role": "system", "content": (
                "I am aware that you are not a financial analyst, far from it."
                "I also know that no serious financial analyst would dare to predict the future or induce people to invest based on their recommendations."
                "However, I want you to gather the opinions of recognized analysts and tell me which 10 companies are presumed to have their stocks on the rise in the next six months."
            )},
            {"role": "user", "content": (
                "Your only response should be a valid Python list, for example ['Figma','Monster','Itaú','Xiaomi', ... 'reflection...'], with no explanations, no additional text."
                "Return them ordered, first the one expected to rise the most and so on."
                "The last item in the array should not be one of the rising companies, but a random suggestion of a good trading practice to improve, considering general investment knowledge, but also practical suggestions for eToro tools."
            )}
        ]

        for attempt in range(1, max_retries + 1):
            try:
                response = client.chat.completions.create(model="gpt-4",
                messages=prompt_messages)
                top10_text = response.choices[0].message.content

                # Intentar interpretar como lista Python
                top10 = ast.literal_eval(top10_text)

                # Validar que sea lista con al menos un string
                if isinstance(top10, list) and all(isinstance(x, str) for x in top10) and len(top10) > 0:
                    return top10
                else:
                    print("⚠️ La respuesta no es una lista válida de strings no vacía.")

            except Exception as e:
                print(f"⚠️ Error al interpretar la respuesta (intento {attempt}): {e}")

            if attempt < max_retries:
                print("🔁 Reintentando...")
                time.sleep(delay)

        # Si todo falla, retornar None
        return None

    # Llamar a la función
    top10 = get_top10_crypto()

    if top10:
        print("💰 Top 10 obtenido: ", top10)
        top10_telegram = "💰 Top 3 obtenido: " + ", ".join(top10[:3])
        df_top10 = pd.DataFrame({top10_header: top10})

        for r in dataframe_to_rows(df_top10, index=False, header=True):
            sheet.append(r)
            
         # --- Función para buscar si ya existe una fila con cierto texto ---
        def find_row_index(sheet, target_text):
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and row[0] == target_text:
                    return i
            return None

        # --- Insertar o sobrescribir Top 10 ---
        top10_row_idx = find_row_index(sheet, top10_header)

        if top10_row_idx:
            for i, r in enumerate(dataframe_to_rows(df_top10, index=False, header=True)):
                for j, value in enumerate(r, start=1):
                    sheet.cell(row=top10_row_idx + i, column=j, value=value)
        else:
            for r in dataframe_to_rows(df_top10, index=False, header=True):
                sheet.append(r)
    else:
        print("❌ No se pudo obtener un Top 10 válido.")

if bajas:
     # Crear hoja "top10"
    top10_bajas_header = "Top 10 a bajo costo"
    if top10_bajas_header in book.sheetnames:
        del book[top10_bajas_header]

    sheet = book.create_sheet(top10_bajas_header)

    def get_top10_bajas_acciones(max_retries=3, delay=2):
        prompt_messages = [
            {"role": "system", "content": (
                "I am aware that you are not a financial analyst, far from it."
                "I also know that no serious financial analyst would dare to predict the future or induce people to invest based on their recommendations."
                "However, I want you to gather the opinions of the most prestigious analysts and tell me which 10 companies that start at a very low actions price are presumed to have potential to rise in the next six months."
            )},
            {"role": "user", "content": (
                "Your only response should be a valid Python list, for example ['Figma','Monster','Itaú','Xiaomi', ... 'reflection...'], with no explanations, no additional text."
                "The last item in the array should not be one of the rising companies, but a random suggestion of a good trading practice to improve, considering general investment knowledge, but also practical suggestions for eToro tools."
            )}
        ]

        for attempt in range(1, max_retries + 1):
            try:
                response = client.chat.completions.create(model="gpt-4",
                messages=prompt_messages)
                top10_bajas_text = response.choices[0].message.content

                # Intentar interpretar como lista Python
                top10_bajas = ast.literal_eval(top10_bajas_text)

                # Validar que sea lista con al menos un string
                if isinstance(top10_bajas, list) and all(isinstance(x, str) for x in top10_bajas) and len(top10_bajas) > 0:
                    return top10_bajas
                else:
                    print("⚠️ La respuesta no es una lista válida de strings no vacía.")

            except Exception as e:
                print(f"⚠️ Error al interpretar la respuesta (intento {attempt}): {e}")

            if attempt < max_retries:
                print("🔁 Reintentando...")
                time.sleep(delay)

        # Si todo falla, retornar None
        return None

    # Llamar a la función
    top10_bajas = get_top10_bajas_acciones()

    if top10_bajas:
        print("💰 Top 10 de bajas obtenido:", top10_bajas)
        bajas_telegram = "💰 Top 3 a bajo costo con potencial: " + ", ".join(top10_bajas[:3])
        df_top10_bajas = pd.DataFrame({top10_bajas_header: top10_bajas})

        for r in dataframe_to_rows(df_top10_bajas, index=False, header=True):
            sheet.append(r)
        
         # --- Función para buscar si ya existe una fila con cierto texto ---
        def find_row_index(sheet, target_text):
            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if row and row[0] == target_text:
                    return i
            return None

        # --- Insertar o sobrescribir Top 10 ---
        top10_bajas_row_idx = find_row_index(sheet, top10_bajas_header)

        if top10_bajas_row_idx:
            for i, r in enumerate(dataframe_to_rows(df_top10_bajas, index=False, header=True)):
                for j, value in enumerate(r, start=1):
                    sheet.cell(row=top10_bajas_row_idx + i, column=j, value=value)
        else:
            for r in dataframe_to_rows(df_top10_bajas, index=False, header=True):
                sheet.append(r)
    else:
        print("❌ No se pudo obtener un Top 10 válido.")

## ----- POLYGON HELPERS -----

# Rate limiter simple (usa el que te pasé antes si querés)
class PolygonForbidden(Exception):
    """HTTP 403: tu plan no está habilitado para el endpoint/dato solicitado."""

def _http_get_json(url: str, params: dict) -> dict:
    """GET con manejo de errores y retorno JSON."""
    r = requests.get(url, params=params, timeout=15)
    try:
        r.raise_for_status()
        return r.json()
    except requests.HTTPError as e:
        # mensaje útil si viene JSON
        try:
            msg = r.json().get("message") or r.text[:400]
        except Exception:
            msg = r.text[:400]
        if r.status_code == 403:
            raise PolygonForbidden(f"403 Forbidden: {msg}") from e
        raise RuntimeError(f"HTTP {r.status_code}: {msg}") from e
    except requests.RequestException as e:
        raise RuntimeError(f"Error de red: {e}") from e
class RateLimiter:
    def __init__(self, max_calls: int, period: float = 60.0):
        self.max_calls = max_calls
        self.period = period
        self.calls: List[float] = []

    def acquire(self) -> None:
        now = time.monotonic()
        self.calls = [t for t in self.calls if now - t < self.period]
        if len(self.calls) >= self.max_calls:
            sleep_for = self.period - (now - self.calls[0])
            if sleep_for > 0:
                time.sleep(sleep_for)
        self.calls.append(time.monotonic())

# Configurá esto a tu plan (ej.: 5 rpm plan free)
GLOBAL_RL = RateLimiter(max_calls=5, period=60.0)

def _chunks(seq: Iterable[str], size: int) -> Iterable[List[str]]:
    buff: List[str] = []
    for x in seq:
        buff.append(x)
        if len(buff) == size:
            yield buff
            buff = []
    if buff:
        yield buff

def _safe_get(d: dict, path: List[str], default=None):
    cur = d
    for k in path:
        if not isinstance(cur, dict) or k not in cur:
            return default
        cur = cur[k]
    return cur

def _company_name(ticker: str, api_key: str, cache: Dict[str, str]) -> str:
    """
    Busca el nombre oficial de la compañía para un ticker.
    Si falla o no hay acceso, devuelve el mismo ticker.
    """
    if ticker in cache:
        return cache[ticker]
    try:
        data = _http_get_json(REF_TICKER_URL_TMPL.format(ticker=ticker), {"apiKey": api_key})
        name = _safe_get(data, ["results", "name"])
        cache[ticker] = (name or "").strip() or ticker
    except Exception:
        cache[ticker] = ticker
    return cache[ticker]

def fetch_quotes_polygon(
    symbols_csv: str,
    api_key: str,
    date: Optional[str] = None,
    *,
    snapshot_chunk_size: int = 50,  # ajustá si tu endpoint limita cantidad por request
) -> dict:
    """
    - Si date (YYYY-MM-DD): histórico -> 1 llamada por ticker (no hay multi-ticker); usa rate limit.
    - Si no hay date: intenta snapshot multi-ticker (chunkeado). Si 403 -> fallback a prev por ticker con rate limit.
    Retorna {"mode": "snapshot"|"prev"|"historical", "data": ...}
    """
    symbols: List[str] = [s.strip().upper() for s in symbols_csv.split(",") if s.strip()]
    if not symbols:
        raise ValueError("Debes pasar al menos un símbolo separado por coma.")

    # ---- Caso histórico (1 a 1) ----
    if date:
        results: Dict[str, Any] = {}
        for t in symbols:
            GLOBAL_RL.acquire()
            hist = _http_get_json(HIST_URL_TMPL.format(ticker=t, date=date),
                                  {"adjusted": "true", "apiKey": api_key})
            results[t] = hist
        return {"mode": "historical", "data": results}

    # ---- Caso actual: snapshot multi-ticker (en chunks) ----
    try:
        all_snaps = []
        for batch in _chunks(symbols, snapshot_chunk_size):
            GLOBAL_RL.acquire()
            snap = _http_get_json(SNAPSHOT_URL, {"tickers": ",".join(batch), "apiKey": api_key})
            all_snaps.append(snap)
        # combiná según tu estructura real de respuesta
        combined = _merge_snapshot_payloads(all_snaps)
        return {"mode": "snapshot", "data": combined}
    except PolygonForbidden:
        # ---- Fallback: prev por ticker (1 a 1) ----
        results: Dict[str, Any] = {}
        for t in symbols:
            GLOBAL_RL.acquire()
            prev = _http_get_json(PREV_URL_TMPL.format(ticker=t),
                                  {"adjusted": "true", "apiKey": api_key})
            results[t] = prev
        return {"mode": "prev", "data": results}

def _merge_snapshot_payloads(payloads: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Une varias respuestas de snapshot multi-ticker en una sola estructura.
    Ajustá esto a la forma real que devuelva tu endpoint.
    """
    out: Dict[str, Any] = {"tickers": []}
    for p in payloads:
        tickers = p.get("tickers") or p.get("results") or []
        if isinstance(tickers, dict):
            # Si viniera en dict, normalizá a lista
            tickers = [tickers]
        out["tickers"].extend(tickers)
    return out

def _extract_prices_from_snapshot(snapshot_payload: Dict[str, Any]) -> Dict[str, float]:
    """
    Ajustá las rutas según tu snapshot real. Ejemplos comunes:
    - lastTrade.p (precio del último trade)
    - day.c (close del día en curso)
    """
    prices: Dict[str, float] = {}
    for item in snapshot_payload.get("tickers", []):
        t = item.get("ticker")
        # Ejemplo de prioridades: lastTrade -> day.close -> prevDay.close
        price = (
            (item.get("lastTrade") or {}).get("p") or
            (item.get("day") or {}).get("c") or
            (item.get("prevDay") or {}).get("c")
        )
        if t and isinstance(price, (int, float)):
            prices[t] = float(price)
    return prices

def _extract_prices_from_prev(prev_payloads: Dict[str, Any]) -> Dict[str, float]:
    """
    Estructura típica de prev: por ticker, results -> [ { c: close } ] ó results: { c: ... }
    Ajustá según tu respuesta.
    """
    prices: Dict[str, float] = {}
    for t, data in prev_payloads.items():
        # Intenta varias formas comunes
        r = data.get("results")
        if isinstance(r, list) and r:
            c = r[0].get("c")
        elif isinstance(r, dict):
            c = r.get("c")
        else:
            c = data.get("close")  # por si viene directo
        if isinstance(c, (int, float)):
            prices[t] = float(c)
    return prices

def print_quotes_polygon(
    symbols_csv: str,
    api_key: str,
    date: Optional[str] = None,
    txt: bool = False
):
    """
    Devuelve (o imprime) { "Nombre Compañía": precio }.
    """
    resp = fetch_quotes_polygon(symbols_csv, api_key, date)
    mode = resp.get("mode")
    data = resp.get("data", {})

    if mode == "snapshot":
        prices_by_ticker = _extract_prices_from_snapshot(data)
    else:  # "prev" o "historical"
        prices_by_ticker = _extract_prices_from_prev(data)

    mapped: Dict[str, float] = {}
    for t, price in prices_by_ticker.items():
        name = _company_name(t, api_key, name_cache) if txt else t  # usa caché
        mapped[name] = price

    # Guardar ambas versiones
    global current_investment
    current_investment = mapped  # queda como dict, usable en send_telegram_table
    current_investment_json = json.dumps(mapped, indent=2, ensure_ascii=False)

    return current_investment_json if txt else mapped

def get_price_by_action(symbol: str, date: Optional[str] = args.date) -> dict:
    ticker = print_quotes_polygon(symbol.upper(), API_KEY_POLYGON, date)
    first_value = list(ticker.values())[0]
    return first_value

## ----- POLYGON FINISHED -----

# --- Tu loop principal, usando las helpers ---
def process_actions():
    for symbol in TICKETS_SYMBOLS:
        if getattr(args, symbol.lower(), None):
            try:
                monto = montos_usd[symbol]
                if not monto:
                    monto = 0  # si no hay monto, no seguimos

                date_company = dates[symbol] or date.today().strftime("%Y-%m-%d")
                precio_historico = get_price_by_action(symbol, date_company)

                if not precio_historico:
                    print(f"⚠️ No se pudo obtener el precio histórico para {symbol} en {date_company}")
                    continue  # no seguimos si no hay precio

                cantidad = round(monto / precio_historico, 8)
                precio_compra = round(monto, 2)
                valor_actual = round(precio_historico, 2)
                wallet = wallets[symbol] or ""

                fila = {
                    "Fecha de Compra": date_company,
                    "Compañía": symbol,
                    "Cantidad": cantidad,
                    "Precio Compra (USD)": precio_compra,
                    "Valor Actual (USD)": valor_actual,
                    "Plataforma": wallet,
                }

                currency_updates.append(symbol)
                df = pd.DataFrame([fila])

                # Crear solapa si no existe
                if symbol in book.sheetnames:
                    sheet = book[symbol]
                else:
                    sheet = book.create_sheet(symbol)
                    # encabezados
                    for r in dataframe_to_rows(pd.DataFrame(columns=HEADERS), index=False, header=True):
                        sheet.append(r)
                    # estilos de encabezado
                    for col_index, header in enumerate(HEADERS, start=1):
                        cell = sheet.cell(row=1, column=col_index)
                        cell.fill = PatternFill(
                            start_color=COLUMN_ACTION_COLORS[symbol],
                            end_color=COLUMN_ACTION_COLORS[symbol],
                            fill_type="solid"
                        )
                        cell.font = Font(color="FFFFFF", bold=True)

                # Si no hay cantidad (puede pasar por alguna división), no tiene sentido registrar
                if cantidad:
                    # Agregar fila inmediatamente después de la última con datos reales
                    write_df_after_last(sheet, df)
                    print(f"✔ Entrada registrada para {symbol}")

                # Calcular totales
                headers = [cell.value for cell in sheet[1]]
                try:
                    cantidad_idx = headers.index("Cantidad") + 1
                    precio_idx = headers.index("Precio Compra (USD)") + 1
                except ValueError:
                    continue

                resumen_row = 2
                col_tot_cantidad = 7  # columna G
                col_tot_precio = 8    # columna H

                # Etiquetas Totales
                cellTotal = sheet.cell(row=1, column=col_tot_cantidad, value="Total de acciones")
                cellPrecio = sheet.cell(row=1, column=col_tot_precio, value="Total USD")
                bgCell = PatternFill(
                    start_color=COLUMN_ACTION_COLORS[symbol],
                    end_color=COLUMN_ACTION_COLORS[symbol],
                    fill_type="solid"
                )
                colorCell = Font(color="FFFFFF", bold=True)
                cellTotal.fill = cellPrecio.fill = bgCell
                cellTotal.font = cellPrecio.font = colorCell

                # Usa la última fila con datos reales, no max_row (que puede estar inflado por formatos)
                ultima_dato = last_data_row(sheet, header_row=1, data_cols=DATA_COLS)

                suma_cantidad = (
                    f"=SUM({get_column_letter(cantidad_idx)}2:{get_column_letter(cantidad_idx)}{ultima_dato})"
                    if cantidad_idx else "0"
                )
                suma_precio = (
                    f"=SUM({get_column_letter(precio_idx)}2:{get_column_letter(precio_idx)}{ultima_dato})"
                    if precio_idx else "0"
                )
                sheet.cell(row=resumen_row, column=col_tot_cantidad, value=suma_cantidad)
                sheet.cell(row=resumen_row, column=col_tot_precio, value=suma_precio)

            except Exception as e:
                print(f"⚠️ Error procesando {symbol}: {e}")
                continue
                
def get_totales():
    # Totales
    global actions_sheet_name
    if actions_sheet_name in book.sheetnames:
        del book[actions_sheet_name]

    sheet = book.create_sheet(actions_sheet_name)
    # Encabezados + columnas de totales
    all_prices = print_quotes_polygon(TICKETS, API_KEY_POLYGON, args.date, False)   

    headers = [
        "Compañía",
        "Cantidad Total",
        "Precio Compra Total (USD)",
        "Cantidad Actual",
        "Ganancia/Perdida (USD)",
        "Total Compra",   # suma de toda la col "Precio Compra Total (USD)"
        "Total Actual",   # suma de toda la col "Cantidad Actual"
        "Total Ganancia"  # suma de toda la col "Ganancia/Perdida (USD)"
    ]
    sheet.append(headers)

    # Estilo de encabezados
    for col_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col_index)
        cell.fill = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
        cell.font = Font(color="FFFFFF", bold=True)

    # Reservamos la fila 2 para las sumas (debajo del encabezado)
    sheet.insert_rows(2)

    # Insertar nueva fila para cada cripto con datos (comienzan en fila 3)
    for symbol in TICKETS_SYMBOLS:

        cantidad_total = 0.0
        precio_total = 0.0

        if symbol in book.sheetnames:
            sheet_symbol = book[symbol]
            for row in sheet_symbol.iter_rows(min_row=2, values_only=True):
                try:
                    # Índices según tus HEADERS en cada hoja de cripto:
                    # 0 Fecha de Compra, 1 Cripto, 2 Cantidad, 3 Precio Compra (USD), 4 Valor Actual (USD), ...
                    cantidad_total += float(row[2]) if row[2] not in (None, "", "N/A") else 0.0
                    precio_total   += float(row[3]) if row[3] not in (None, "", "N/A") else 0.0
                except (ValueError, IndexError) as e:
                    print(f"⚠️ Error procesando fila en {symbol}: {row} — {e}")
                    # continue

        precio_actual = 0.0
        precio_actual = all_prices.get(symbol.strip().upper())
        cantidad_actual = round(cantidad_total * precio_actual, 2) if cantidad_total and precio_actual else ""
        cantidad_actual = round(cantidad_total * precio_actual, 2) if cantidad_total and precio_actual else ""
        ganancia_perdida = round(cantidad_actual - precio_total, 2) if cantidad_actual != "" else ""

        sheet.append([symbol, cantidad_total, precio_total, cantidad_actual, ganancia_perdida, "", "", ""])

    # Ahora que ya están las filas, escribimos las SUMAS en la fila 2
    last_row = sheet.max_row

    # Columnas (1-indexed) según 'headers'
    col_precio_compra = headers.index("Precio Compra Total (USD)") + 1  # C = 3
    col_cantidad_actual = headers.index("Cantidad Actual") + 1           # D = 4
    col_ganancia = headers.index("Ganancia/Perdida (USD)") + 1          # E = 5
    col_total_compra = headers.index("Total Compra") + 1                # F = 6
    col_total_actual = headers.index("Total Actual") + 1                # G = 7
    col_total_ganancia = headers.index("Total Ganancia") + 1            # H = 8

    # Si hay al menos una fila de datos (fila 3 en adelante), ponemos las fórmulas; si no, "0"
    if last_row >= 3:
        rango_compra   = f"{get_column_letter(col_precio_compra)}2:{get_column_letter(col_precio_compra)}{last_row}"
        rango_actual   = f"{get_column_letter(col_cantidad_actual)}2:{get_column_letter(col_cantidad_actual)}{last_row}"
        rango_ganancia = f"{get_column_letter(col_ganancia)}2:{get_column_letter(col_ganancia)}{last_row}"

        sheet.cell(row=2, column=col_total_compra,   value=f"=SUM({rango_compra})")
        sheet.cell(row=2, column=col_total_actual,   value=f"=SUM({rango_actual})")
        sheet.cell(row=2, column=col_total_ganancia, value=f"=SUM({rango_ganancia})")
    else:
        sheet.cell(row=2, column=col_total_compra,   value="0")
        sheet.cell(row=2, column=col_total_actual,   value="0")
        sheet.cell(row=2, column=col_total_ganancia, value="0")

    for col in (col_total_compra, col_total_actual, col_total_ganancia):
        c = sheet.cell(row=2, column=col)
        c.font = Font(bold=True)

def _read_sheet_values(excel_path: str, sheet_name: str) -> pd.DataFrame:

    wb = load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.values)
    if not rows:
        return pd.DataFrame()
    headers = list(rows[0])
    data = rows[1:]
    df = pd.DataFrame(data, columns=headers)
    return df

def _format_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    """2 dec para USD/totales, 8 dec para cantidades; resto texto."""
    out = df.copy()
    usd_like = { "Cripto",
        "Cantidad Total",
        "Precio Compra Total (USD)",
        "Cantidad Actual",
        "Ganancia/Perdida (USD)"}
    for c in out.columns:
        if pd.api.types.is_numeric_dtype(out[c]):
            if "Cantidad" in c and "USD" not in c:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.8f}".replace(",", "X").replace(".", ",").replace("X", "."))
            elif c in usd_like or "USD" in c:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
            else:
                out[c] = out[c].map(lambda x: "" if pd.isna(x) else f"{float(x):,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        else:
            out[c] = out[c].astype(str)
    return out

def _send_html_pre(chat_id: str, title: str, body: str):
    if not TELEGRAM_TOKEN:
        raise RuntimeError("Falta TELEGRAM_TOKEN")
    payload = {
        "chat_id": chat_id or DEFAULT_CHAT_ID,
        "text": f"{html.escape(title)}\n<pre>{html.escape(body)}</pre>",
        "parse_mode": "HTML",
        "disable_web_page_preview": True,
    }
    r = requests.post(TELEGRAM_API, json=payload, timeout=30)
    try:
        r.raise_for_status()
        print("💰 Mensaje enviado a Telegram")
        return r.json()
    except requests.HTTPError as e:
        # mensaje útil si viene JSON
        try:
            msg = r.json().get("message") or r.text[:400]
        except Exception:
            msg = r.text[:400]
        if r.status_code == 403:
            raise PolygonForbidden(f"403 Forbidden: {msg}") from e
        raise RuntimeError(f"HTTP {r.status_code}: {msg}") from e
    except requests.RequestException as e:
        raise RuntimeError(f"Error de red: {e}") from e

def send_telegram_table(
    sheet_name: str,
    *,
    excel_path: str,
    columns: Optional[List[str]] = None,
    top: Optional[int] = None,
    sort_by: Optional[str] = None,
    ascending: bool = False,
    rows_per_msg: int = 12,
    tablefmt: str = "github",
    title: Optional[str] = None,
    chat_id: Optional[str] = None,
) -> None:
    chat_id = chat_id or DEFAULT_CHAT_ID
    if not chat_id:
        raise RuntimeError("Falta DEFAULT_CHAT_ID/TELEGRAM_CHAT_ID")
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No encuentro el Excel en: {excel_path}")

    # 1) Leer hoja
    df = _read_sheet_values(excel_path, sheet_name)

    # 2) Filtrar y ordenar
    if columns:
        cols = [c for c in columns if c in df.columns]
        df = df[cols] if cols else df
    if sort_by and sort_by in df.columns:
        df = df.sort_values(sort_by, ascending=ascending)

    if top is not None:
        df = df.head(top)

    # === NUEVO: calcular totales y diferencia sobre datos numéricos ===
    col_precio = "Precio Compra Total (USD)"
    col_cant   = "Cantidad Actual"

    precio_num = pd.to_numeric(df[col_precio], errors="coerce") if col_precio in df.columns else pd.Series(dtype=float)
    cant_num   = pd.to_numeric(df[col_cant], errors="coerce")   if col_cant   in df.columns else pd.Series(dtype=float)

    total_precio = float(precio_num.sum(skipna=True)) if len(precio_num) else 0.0
    total_cant   = float(cant_num.sum(skipna=True))   if len(cant_num)   else 0.0
    diferencia   = total_cant - total_precio

    # 5) Limpieza final para visual
    df = df.replace({None: pd.NA}).fillna("")

    # 4) Formatear números para la tabla
    df_fmt = _format_numeric_columns(df)
    n = len(df_fmt)
    pages = max(1, math.ceil(n / rows_per_msg)) if rows_per_msg else 1

    for p in range(pages):
        start = p * rows_per_msg
        end = None if not rows_per_msg else start + rows_per_msg
        page_df_fmt = df_fmt.iloc[start:end]  # 👈 usar el slice correcto

        body = tabulate(
            page_df_fmt,
            headers="keys",
            tablefmt=tablefmt,
            numalign="right",
            stralign="left",
            showindex=False
        )

        # === NUEVO: agregar totales al final de la ÚLTIMA página ===
        if p == pages - 1:

            footer_lines = [
                "",
                f"Total Precio Compra: {total_precio:,.0f}",
                f"Total Cantidad Actual: {total_cant:,.0f}",
                f"✅ Ganancia: {diferencia:,.0f}" if (diferencia > 0) else f"❌ Pérdida: {diferencia:,.0f}",
            ]
            body = body + "\n" + "\n".join(footer_lines) + "\n\n"
            
        if top10_telegram != "":
            body = body + "\n" + top10_telegram
        if bajas_telegram != "":
            body = body + "\n" + bajas_telegram

        # Recorte ultraseguro si excede límites
        if len(body) > 3900:
            body = "\n".join(line[:120] for line in body.splitlines())

        page_title = f"{title or sheet_name} ({p+1}/{pages})" if pages > 1 else (title or sheet_name)
        _send_html_pre(chat_id, page_title, body)        
        
def get_telegram():
    headers_list = [
        "Compañía",
        "Cantidad Total",
        "Precio Compra Total (USD)",
        "Cantidad Actual",
        "Ganancia/Perdida (USD)"
    ]

    send_telegram_table(
        sheet_name=actions_sheet_name,
        excel_path=EXCEL_FILE,        # usa tu constante existente
        columns=headers_list,
        top=20,
        sort_by="Total Ganancia",     # opcional: ordena por ganancia
        ascending=False,
        rows_per_msg=12,
        tablefmt="github",
        title=actions_sheet_name,
        chat_id=DEFAULT_CHAT_ID
    )
    
if __name__ == "__main__":
    if not API_KEY_POLYGON or not TICKETS:
        print(
            file=sys.stderr
        )
        sys.exit(2)

    try:
        # Usa exactamente las variables que definiste
        if telegram:
            print_quotes_polygon(TICKETS, API_KEY_POLYGON, args.date, True)
            get_telegram()
            
        process_actions()
        get_totales()
        
        # Guardar archivo
        book.save(EXCEL_FILE)
        print(f"💰 Archivo actualizado: {EXCEL_FILE}")
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)